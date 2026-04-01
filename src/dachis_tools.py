'''
    This file has multiple functions.
    See below for each one.
    Pretty much everything is written by ChatGPT.
    This is supplementary code only.
    
    Written by ChatGPT for Dachi Dzeria — extended from base version.
    I do not claim any ownership. This is only a tool for my projects.    
        - Dachi Dzeria, 27/10/2025
'''
import os
import subprocess
import platform
from datetime import datetime

def get_attributes(obj, max_depth=2, _depth=0, _visited=None):
    """
    This is a function written to get the subfunctions (attributes of attributes)
    of any Python object. Exports the results into a .txt file on the Desktop.
    
    Recursively introspects an object's attributes using dir(),
    listing sub-attributes up to `max_depth` levels deep.
    Filters out dunder methods/attributes (anything starting and ending with __).
    
    Parameters
    ----------
    obj : any
        The object to introspect.
    max_depth : int, optional
        Maximum recursion depth (default: 2).
    _depth : int, internal use
        Tracks current recursion depth.
    _visited : set, internal use
        Tracks visited object IDs to avoid infinite loops.
    """
    if _visited is None:
        _visited = set()
    
    obj_type = type(obj)
    obj_name = getattr(obj, "__name__", obj_type.__name__)
    obj_id = id(obj)
    
    if obj_id in _visited:
        return [f"{'    '*_depth}[Already visited: {obj_name}]"]
    
    _visited.add(obj_id)
    output_lines = []
    indent = "    " * _depth
    
    if _depth == 0:
        output_lines += [
            f"Object: {obj_name}",
            f"Type: {obj_type}",
            f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "Attributes from dir() (dunder methods excluded):",
            "------------------------------------------------",
        ]
    
    try:
        attributes = [a for a in dir(obj) if not (a.startswith("_") and a.endswith("_"))]
    except Exception as e:
        output_lines.append(f"{indent}[dir() failed: {e}]")
        return output_lines
    
    for attr in attributes:
        try:
            val = getattr(obj, attr)
            val_type = type(val).__name__
            output_lines.append(f"{indent}- {attr} ({val_type})")
            
            if (
                _depth + 1 < max_depth
                and not isinstance(val, (int, float, str, bool, bytes, type(None)))
            ):
                output_lines += get_attributes(val, max_depth, _depth + 1, _visited)
        except Exception as e:
            output_lines.append(f"{indent}- {attr} [Error accessing: {e}]")
    
    # Write only once (at root call)
    if _depth == 0:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        filename = f"{obj_name}_attributes.txt"
        file_path = os.path.join(desktop_path, filename)
        
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))
        
        print(f"PASS: Exported recursive dir() of '{obj_name}' to:\n{file_path}")
        
        # Auto-open file
        try:
            system_name = platform.system()
            if system_name == "Windows":
                os.startfile(file_path)
            elif system_name == "Darwin":
                subprocess.run(["open", file_path])
            elif system_name == "Linux":
                subprocess.run(["xdg-open", file_path])
        except Exception as e:
            print(f"[!] Could not open file automatically: {e}")
        
        return file_path
    
    return output_lines

# %% read_table
# # -*- coding: utf-8 -*-
"""
    Created on Sat Nov  1 17:08:54 2025

    Opens an Excel file, finds a table by name in a given sheet, 
    and converts the table to a pandas DataFrame.
    
    Author: ChatGPT, pretty much
"""

import pandas as pd
from openpyxl import load_workbook

def read_table(file_path, sheet_name, table_name):
    """
    Opens an Excel file, finds a table by name in a given sheet, 
    and converts the table to a pandas DataFrame.

    Parameters
    ----------
    file_path : str
        Path to the Excel file (.xlsx)
    sheet_name : str
        Name of the sheet where the table is located
    table_name : str
        Name of the Excel table to read

    Returns
    -------
    pd.DataFrame
        DataFrame containing the contents of the specified table
    """

    # Load workbook and specified sheet
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]

    # Check if the table exists
    if table_name not in ws._tables:
        available = list(ws._tables.keys())
        raise ValueError(f"Table '{table_name}' not found. Available tables: {available}")

    # Get the table object and its range
    table = ws._tables[table_name]
    ref = table.ref

    # Convert Excel range to a pandas DataFrame
    data = ws[ref]

    # Extract rows as lists of values
    rows = [[cell.value for cell in row] for row in data]
    header = rows[0]
    values = rows[1:]

    # Create DataFrame
    df = pd.DataFrame(values, columns=header)
    return df


# %% console_logger

import sys

class console_logger:
    def __init__(self, folder_path=".", file_name = "."):
        """
        Initialize the logger.
        :param folder_path: directory where the log file will be saved
        """
        self.folder_path = folder_path
        self.original_stdout = sys.stdout
        self.log_file = None
        self._last_flush = None

        # Ensure folder exists
        os.makedirs(self.folder_path, exist_ok=True)

        # Create file with timestamp in name
        self.file_path = os.path.join(self.folder_path, f"{file_name}_log.txt")

    def start(self):
        """Redirect stdout to the log file."""
        self.log_file = open(self.file_path, "a")
        self._last_flush = datetime.now()
        sys.stdout = self

    def stop(self):
        """Restore stdout and close the log file."""
        if self.log_file:
            sys.stdout = self.original_stdout
            self.log_file.flush()
            self.log_file.close()
            self.log_file = None
            print(f"PASS: Log saved successfully at: {self.file_path}")

    def write(self, message):
        if message.strip():  # Skip empty messages
            now = datetime.now()
            timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
            formatted_message = f"[{timestamp}] {message}"
            # Ensure a single newline at the end
            if not formatted_message.endswith("\n"):
                formatted_message += "\n"
            self.original_stdout.write(formatted_message)
            self.log_file.write(formatted_message)
            # Flush at most once every 5 seconds
            if (now - self._last_flush).total_seconds() >= 5:
                self.log_file.flush()
                self._last_flush = now

    def flush(self):
        """Flush both stdout and the log file, and reset the flush timer."""
        self.original_stdout.flush()
        if self.log_file:
            self.log_file.flush()
            self._last_flush = datetime.now()

# %% new_folder_and_file

import shutil
def new_folder_and_file(file_path, nickname = "", copy = False):
    """
    Takes a file path, creates a new folder in the same directory,
    copies the file into it (with a timestamp in the name),
    and returns both the new file path and folder path.
    Args:
        file_path (str): Path to the original file.
    Returns:
        tuple[str, str]: (new_file_path, new_folder_path)
    """
    # Ensure the file exists
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    # Extract components
    parent_dir = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    name, ext = os.path.splitext(file_name)
    # Create timestamp
    timestamp = datetime.now().strftime("%M%S")
    # Create new folder (based on original file)
    new_folder = os.path.join(parent_dir, f"{nickname}_{timestamp}")
    os.makedirs(new_folder, exist_ok=True)
    # Build new file name with timestamp
    if copy:
        new_file_name = f"{name}_{ext}"
    else:
        new_file_name = f"{name}"
    
    new_file_path = os.path.join(new_folder, new_file_name)
    
    if copy: # Copy file with new name
        shutil.copy2(file_path, new_file_path)
    
    print("PASS: Folder Created")
    
    return new_file_path, new_folder

# %% Results analysis

import re

def find_and_publish_results(paths: list[str], phrases: list[str], file_types: list[str], columns: list[str] = None):
    """
    Runs the phrase-number extraction for multiple base paths and file types.
    Creates a separate 'results.xlsx' in each path.
    
    Args:
        paths: List of folder paths to search.
        phrases: List of phrases to look for.
        file_types: List of file extensions to search (e.g., [".txt", ".log", ".md"]).
        columns: Optional list of column names to rename the phrases.
    
    Returns:
        A dictionary mapping each path to its resulting DataFrame.
    """
    results = {}
    for base_path in paths:
        print(f"Processing path: {base_path}")
        data = []
        # Compile regex patterns for all phrases
        patterns = {
            phrase: re.compile(re.escape(phrase) + r'\s*([-+]?\d*\.?\d+)', re.IGNORECASE)
            for phrase in phrases
        }
        for root, _, files in os.walk(base_path):
            for file in files:
                # Check if file matches any of the specified file types
                if any(file.lower().endswith(ext.lower()) for ext in file_types):
                    file_path = os.path.join(root, file)
                    row_data = {"File": os.path.relpath(file_path, base_path)}
                    try:
                        with open(file_path, "r", encoding="utf-8") as f:
                            content = f.read()
                        for phrase, pattern in patterns.items():
                            matches = pattern.findall(content)
                            row_data[phrase] = ", ".join(matches) if matches else None
                        data.append(row_data)
                    except (UnicodeDecodeError, OSError) as e:
                        print(f"Could not read {file_path}: {e}")
        # Convert to DataFrame
        df = pd.DataFrame(data)
        # Rename columns if 'columns' provided
        if columns is not None and len(columns) == len(phrases):
            rename_map = dict(zip(phrases, columns))
            df = df.rename(columns=rename_map)
            cols = ["File"] + columns
        else:
            cols = ["File"] + phrases
        df = df.reindex(columns=cols)
        # Save Excel file in the base path
        output_file = os.path.join(base_path, "results.xlsx")
        df.to_excel(output_file, index=False)
        print(f"PASS: Results saved to: {output_file}")
        # Store the DataFrame in the dictionary
        results[base_path] = df
    return results

# %% move_file

from pathlib import Path

def move_file(source, destination):
    """Copies file to folder (use shutil.move to actually move)"""
    source = Path(source)
    destination = Path(destination) / source.name
    shutil.copy2(source, destination)
    print(f"PASS: {source} moved to {destination}")

# %% Find results (without publishing)

def find_results(paths: list[str], phrases: list[str], file_types: list[str], 
                 columns: list[str] = None):
    """
    Extracts phrase-number pairs from files, grouping by subfolder.
    
    Args:
        paths: List of folder paths to search (each returns its own DataFrame).
        phrases: List of phrases to look for.
        file_types: List of file extensions to search (e.g., [".txt", ".log", ".out"]).
        columns: Optional list of column names to rename the phrases.
    
    Returns:
        A tuple of DataFrames (one per path), where each row is a subfolder.
        Empty DataFrames are returned for paths with no data.
    """
    results = []
    
    # Process each path separately
    for base_path in paths:
        print(f"Processing path: {base_path}")
        
        # Check if path exists
        if not os.path.exists(base_path):
            print(f"Warning: Path does not exist, returning empty DataFrame: {base_path}")
            # Create empty DataFrame with correct columns
            if columns is not None and len(columns) == len(phrases):
                empty_df = pd.DataFrame(columns=["Subfolder"] + columns)
            else:
                empty_df = pd.DataFrame(columns=["Subfolder"] + phrases)
            results.append(empty_df)
            continue
        
        # Compile regex patterns for each phrase
        # Updated pattern to handle scientific notation (e.g., 3.4e-4, 1.2E+5)
        patterns = {
            phrase: re.compile(re.escape(phrase) + r'\s*([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)', re.IGNORECASE)
            for phrase in phrases
        }
        
        # Get all immediate subfolders
        try:
            subfolders = [f for f in os.listdir(base_path) 
                         if os.path.isdir(os.path.join(base_path, f))]
        except OSError as e:
            print(f"Could not access {base_path}: {e}, returning empty DataFrame")
            # Create empty DataFrame with correct columns
            if columns is not None and len(columns) == len(phrases):
                empty_df = pd.DataFrame(columns=["Subfolder"] + columns)
            else:
                empty_df = pd.DataFrame(columns=["Subfolder"] + phrases)
            results.append(empty_df)
            continue
        
        if not subfolders:
            print(f"No subfolders found in {base_path}, returning empty DataFrame")
            # Create empty DataFrame with correct columns
            if columns is not None and len(columns) == len(phrases):
                empty_df = pd.DataFrame(columns=["Subfolder"] + columns)
            else:
                empty_df = pd.DataFrame(columns=["Subfolder"] + phrases)
            results.append(empty_df)
            continue
        
        data = []
        
        # Process each subfolder
        for subfolder in subfolders:
            subfolder_path = os.path.join(base_path, subfolder)
            row_data = {"Subfolder": subfolder}
            
            # Collect all matches from all files in this subfolder
            all_matches = {phrase: [] for phrase in phrases}
            
            # Walk through the subfolder recursively
            for root, _, files in os.walk(subfolder_path):
                for file in files:
                    # Check if file extension matches
                    if any(file.lower().endswith(ext.lower()) for ext in file_types):
                        file_path = os.path.join(root, file)
                        
                        try:
                            # Read file content
                            with open(file_path, "r", encoding="utf-8") as f:
                                content = f.read()
                            
                            # Search for each phrase
                            for phrase, pattern in patterns.items():
                                matches = pattern.findall(content)
                                all_matches[phrase].extend(matches)
                                
                        except (UnicodeDecodeError, OSError) as e:
                            print(f"Could not read {file_path}: {e}")
            
            # Store results for this subfolder
            for phrase in phrases:
                matches = all_matches[phrase]
                row_data[phrase] = ", ".join(matches) if matches else None
            
            data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Rename columns if custom names provided
        if columns is not None and len(columns) == len(phrases):
            rename_map = dict(zip(phrases, columns))
            df = df.rename(columns=rename_map)
            cols = ["Subfolder"] + columns
        else:
            cols = ["Subfolder"] + phrases
        
        df = df.reindex(columns=cols)
        
        results.append(df)
        print(f"Processed {len(df)} subfolders from {base_path}")
    
    return tuple(results)


# %% Find one result

def find_result(path: str, phrase: str) -> float:
    """
    Scans all .txt files in a path and returns the last numerical value 
    found after the specified phrase.
    """
    pattern = re.compile(re.escape(phrase) + r'\s*([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)')

    # Sorting files ensures deterministic ordering.
    # reverse=True processes files from Z to A.
    files = sorted(os.listdir(path), reverse=True)

    for file in files:
        if file.endswith(".txt"):
            file_path = os.path.join(path, file)
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
                    
                    # finditer returns an iterator of all non-overlapping matches
                    matches = list(pattern.finditer(content))
                    
                    if matches:
                        # Index [-1] grabs the final match in the document
                        return float(matches[-1].group(1))
            except (UnicodeDecodeError, OSError):
                continue
                
    return None


# %% Get cell count

def find_delimited(folder_path, phrase):
    return sum(int(next((line.split()[-1] for line in p.open() if phrase in line), 0)) for p in Path(folder_path).glob("*.txt"))
    
# %% Saves all inputs of a function in a .txt file

def save_inputs(folder_path, filename, inputs):
    """
    Saves a dictionary of inputs to a text file in the specified folder.
    
    Args:
        folder_path (str): Directory to save the file.
        filename (str): Name of the file (e.g., 'trial_001.txt').
        inputs (dict): Dictionary of function arguments.
    """
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    
    with open(file_path, "w") as f:
        # Option 1: Clean, readable Key = Value format
        for key, value in inputs.items():
            f.write(f"{key} = {value}\n")
        
        # Option 2: JSON format (better for re-loading later)
        # json.dump(inputs, f, indent=4)

    return file_path
# %% Get last

def get_last(paths: list[str], file_types: list[str], name_length: int = None):
    """
    Extracts the last value from files in each subfolder.
    Files with the same prefix (first N characters) are grouped into the same column.
    
    Args:
        paths: List of folder paths to search (each returns its own DataFrame).
        file_types: List of file extensions to search (e.g., [".txt", ".log", ".out"]).
        name_length: Optional number of characters from the start of the file name to use as column name.
                     Files with the same prefix go to the same column. If None, uses full file name.
    
    Returns:
        A tuple of DataFrames (one per path), where each row is a subfolder
        and each column represents files with the same prefix.
    """
    results = []
    
    # Process each path separately
    for base_path in paths:
        print(f"Processing path: {base_path}")
        
        # Check if path exists
        if not os.path.exists(base_path):
            print(f"Warning: Path does not exist, returning empty DataFrame: {base_path}")
            empty_df = pd.DataFrame(columns=["Subfolder"])
            results.append(empty_df)
            continue
        
        # Get all immediate subfolders
        try:
            subfolders = [f for f in os.listdir(base_path) 
                         if os.path.isdir(os.path.join(base_path, f))]
        except OSError as e:
            print(f"Could not access {base_path}: {e}, returning empty DataFrame")
            empty_df = pd.DataFrame(columns=["Subfolder"])
            results.append(empty_df)
            continue
        
        if not subfolders:
            print(f"No subfolders found in {base_path}, returning empty DataFrame")
            empty_df = pd.DataFrame(columns=["Subfolder"])
            results.append(empty_df)
            continue
        
        # First pass: collect all unique column names (prefixes) across all subfolders
        all_col_names = set()
        for subfolder in subfolders:
            subfolder_path = os.path.join(base_path, subfolder)
            for root, _, files in os.walk(subfolder_path):
                for file in files:
                    if any(file.lower().endswith(ext.lower()) for ext in file_types):
                        col_name = file[:name_length] if name_length is not None else file
                        all_col_names.add(col_name)
        
        # Sort column names for consistent ordering
        all_col_names = sorted(all_col_names)
        
        data = []
        
        # Process each subfolder
        for subfolder in subfolders:
            subfolder_path = os.path.join(base_path, subfolder)
            row_data = {"Subfolder": subfolder}
            
            # Process each column (prefix)
            for col_name in all_col_names:
                last_value = None
                
                # Search for any file matching this prefix in the subfolder
                for root, _, files in os.walk(subfolder_path):
                    for file in files:
                        if any(file.lower().endswith(ext.lower()) for ext in file_types):
                            file_prefix = file[:name_length] if name_length is not None else file
                            
                            if file_prefix == col_name:
                                file_path = os.path.join(root, file)
                                
                                try:
                                    # Read file content
                                    with open(file_path, "r", encoding="utf-8") as f:
                                        lines = f.readlines()
                                    
                                    # Find last non-empty line
                                    for i in range(len(lines) - 1, -1, -1):
                                        if lines[i].strip():
                                            last_line = lines[i].strip()
                                            # Extract all numbers from the last line
                                            numbers = re.findall(r'[-+]?\d*\.?\d+', last_line)
                                            # Take the last number
                                            if numbers:
                                                last_value = numbers[-1]
                                            break
                                    
                                    # Found a match, use it and stop searching
                                    break
                                    
                                except (UnicodeDecodeError, OSError) as e:
                                    print(f"Could not read {file_path}: {e}")
                    
                    if last_value is not None:
                        break  # Found value, move to next column
                
                row_data[col_name] = last_value
            
            data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Build column order
        cols = ["Subfolder"] + list(all_col_names)
        df = df.reindex(columns=cols)
        
        results.append(df)
        print(f"Processed {len(df)} subfolders from {base_path}")
    
    return tuple(results)

# %% Publishing the results

def publish(dataframes: list, folder_paths: list[str], file_names: list[str] = None):
    """
    Saves each DataFrame to its corresponding folder path as an Excel file.
    
    Args:
        dataframes: List of DataFrames to save.
        folder_paths: List of folder paths where each DataFrame will be saved.
        file_names: Optional list of file names (without extension). If None, uses "results_1", "results_2", etc.
    
    Returns:
        None
    """
    # Check if lists are the same length
    if len(dataframes) != len(folder_paths):
        print(f"Warning: Number of DataFrames ({len(dataframes)}) doesn't match number of folder paths ({len(folder_paths)})")
        return
    
    # Generate default file names if not provided
    if file_names is None:
        file_names = [f"results_{i+1}" for i in range(len(dataframes))]
    elif len(file_names) != len(dataframes):
        print(f"Warning: Number of file names ({len(file_names)}) doesn't match number of DataFrames ({len(dataframes)})")
        return
    
    # Save each DataFrame to its corresponding path
    for df, folder_path, file_name in zip(dataframes, folder_paths, file_names):
        try:
            # Create folder if it doesn't exist
            os.makedirs(folder_path, exist_ok=True)
            
            # Construct full file path
            full_path = os.path.join(folder_path, f"{file_name}.xlsx")
            
            # Save DataFrame
            df.to_excel(full_path, index=False)
            print(f"Saved to: {full_path}")
        except Exception as e:
            print(f"Error saving to {folder_path}: {e}")
            
# %% make_gifs function

import imageio.v3 as iio
from PIL import Image
import numpy as np

def make_gif(output_path, input_path, duration):
        
    # load the path, sorted
    image_paths = sorted(Path(input_path).glob('*.png'))
    
    # open the first image and get its size
    first_img = Image.open(image_paths[0])
    target_size = first_img.size
    
    normalized_images = []
    
    for path in image_paths:
        img = Image.open(path)
        if img.size != target_size:
            # resize
            img = img.resize(target_size, Image.Resampling.LANCZOS)
        
        # not actually sure what this does
        normalized_images.append(np.array(img.convert("RGB")))
        
    # Write
    iio.imwrite(output_path, normalized_images, duration = 200, loop = 0)
    
    
# %% Plotting coefficients
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

def plot_and_save_transient(file_paths):
    if isinstance(file_paths, (str, Path)):
        file_paths = [file_paths]
        
    saved_paths = []
    
    for p in [Path(f) for f in file_paths]:
        flow_times, coefficients = [], []
        with p.open('r') as file:
            for line in file.readlines()[3:]:
                data = line.strip().split()
                if len(data) >= 3:
                    coefficients.append(float(data[1]))
                    flow_times.append(float(data[2]))
        
        fig, ax = plt.subplots(figsize=(10, 6))
        label_name = p.stem.replace('-rfile', '')
        ax.plot(flow_times, coefficients, label=label_name)
        
        last_x, last_y = flow_times[-1], coefficients[-1]
        ax.scatter([last_x], [last_y], color='red', zorder=5)
        ax.annotate(f'Final: {last_y:.6f}', (last_x, last_y), 
                    textcoords="offset points", xytext=(10, -10), ha='left', va='top',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8))

        ax.set_xlabel('Flow Time (s)')
        ax.set_ylabel('Coefficient')
        ax.set_title(f'Transient Aerodynamic Coefficient: {label_name}')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        plt.tight_layout()

        save_path = p.parent / f"{label_name}_transient.png"
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close(fig) 
        
        # Append the string version of the path to the return list
        saved_paths.append(str(save_path))
        
    return saved_paths

def plot_and_save_steady(file_paths):
    if isinstance(file_paths, (str, Path)):
        file_paths = [file_paths]
        
    saved_paths = []
    
    for p in [Path(f) for f in file_paths]:
        iterations, coefficients = [], []
        with p.open('r') as file:
            for line in file.readlines()[3:]:
                data = line.strip().split()
                if len(data) >= 2:
                    iterations.append(int(data[0]))
                    coefficients.append(float(data[1]))
        
        fig, ax = plt.subplots(figsize=(10, 6))
        label_name = p.stem.replace('-rfile', '')
        ax.plot(iterations, coefficients, label=label_name)
        
        last_x, last_y = iterations[-1], coefficients[-1]
        ax.scatter([last_x], [last_y], color='red', zorder=5)
        ax.annotate(f'Final: {last_y:.6f}', (last_x, last_y), 
                    textcoords="offset points", xytext=(10, -10), ha='left', va='top',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8))

        ax.set_xlabel('Iteration')
        ax.set_ylabel('Coefficient')
        ax.set_title(f'Steady State: {label_name}')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        plt.tight_layout()

        save_path = p.parent / f"{label_name}.png"
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close(fig)
        
        # Append the string version of the path to the return list
        saved_paths.append(str(save_path))
        
    return saved_paths